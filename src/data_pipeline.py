import json
import logging
import re
import os
from pathlib import Path

import openpyxl

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
XLSX_PATH = ROOT / "NUST Bank-Product-Knowledge.xlsx"
FAQ_PATH = ROOT / "faq.json"

SKIP_SHEETS = {"Main", "Rate Sheet July 1 2024", "Sheet1"}
CHUNK_TOKEN_TARGET = 600
OVERLAP_TOKENS = 80


def _clean_text(text: str) -> str:
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    text = text.strip()
    return text


def _row_text(row) -> str:
    parts = []
    for cell in row:
        val = cell.value if hasattr(cell, "value") else cell
        if val is not None:
            s = str(val).strip()
            if s and s.lower() != "main" and len(s) > 2:
                parts.append(s)
    return " ".join(parts)


def _extract_xlsx_pairs(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    pairs = []

    for sheet_name in wb.sheetnames:
        if sheet_name in SKIP_SHEETS:
            continue

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        text_rows = []
        for row in rows:
            combined = " ".join(
                str(c).strip() for c in row
                if c is not None and str(c).strip() and str(c).strip().lower() != "main"
            )
            combined = _clean_text(combined)
            if len(combined) > 10:
                text_rows.append(combined)

        if len(text_rows) < 2:
            continue

        # First row is often the sheet/product title
        product_title = text_rows[0]
        data_rows = text_rows[1:]

        # Pair consecutive rows as Q & A
        i = 0
        while i + 1 < len(data_rows):
            q = data_rows[i]
            a = data_rows[i + 1]
            if len(q) > 10 and len(a) > 10:
                pairs.append({
                    "source": sheet_name,
                    "product": product_title,
                    "question": _clean_text(q),
                    "answer": _clean_text(a),
                })
            i += 2

    logger.info(f"Extracted {len(pairs)} Q&A pairs from XLSX")
    return pairs


def _extract_faq_pairs(path: Path) -> list[dict]:
    with open(path, encoding="utf-8") as f:
        data = json.load(f)

    pairs = []
    for cat in data.get("categories", []):
        category = cat.get("category", "General")
        for item in cat.get("questions", []):
            q = _clean_text(item.get("question", ""))
            a = _clean_text(item.get("answer", ""))
            if q and a:
                pairs.append({
                    "source": "faq",
                    "product": category,
                    "question": q,
                    "answer": a,
                })

    logger.info(f"Extracted {len(pairs)} Q&A pairs from FAQ JSON")
    return pairs


def _approx_tokens(text: str) -> int:
    return len(text.split())


def _chunk_text(text: str, target: int = CHUNK_TOKEN_TARGET, overlap: int = OVERLAP_TOKENS) -> list[str]:
    words = text.split()
    chunks = []
    start = 0
    while start < len(words):
        end = min(start + target, len(words))
        chunk = " ".join(words[start:end])
        chunks.append(chunk)
        if end == len(words):
            break
        start = end - overlap
    return chunks


def build_chunks(pairs: list[dict]) -> list[dict]:
    chunks = []
    for pair in pairs:
        block = f"Product: {pair['product']}\nQ: {pair['question']}\nA: {pair['answer']}"
        block = _clean_text(block)
        if _approx_tokens(block) <= CHUNK_TOKEN_TARGET:
            chunks.append({
                "text": block,
                "source": pair["source"],
                "product": pair["product"],
            })
        else:
            for chunk in _chunk_text(block):
                chunks.append({
                    "text": chunk,
                    "source": pair["source"],
                    "product": pair["product"],
                })
    logger.info(f"Built {len(chunks)} chunks total")
    return chunks


def run_pipeline() -> list[dict]:
    DATA_DIR.mkdir(exist_ok=True)

    pairs = []
    if XLSX_PATH.exists():
        pairs.extend(_extract_xlsx_pairs(XLSX_PATH))
    else:
        logger.warning(f"XLSX not found at {XLSX_PATH}")

    if FAQ_PATH.exists():
        pairs.extend(_extract_faq_pairs(FAQ_PATH))
    else:
        logger.warning(f"FAQ JSON not found at {FAQ_PATH}")

    chunks = build_chunks(pairs)

    output_path = DATA_DIR / "chunks.json"
    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(chunks, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved {len(chunks)} chunks to {output_path}")

    pairs_path = DATA_DIR / "qa_pairs.json"
    with open(pairs_path, "w", encoding="utf-8") as f:
        json.dump(pairs, f, ensure_ascii=False, indent=2)
    logger.info(f"Saved {len(pairs)} Q&A pairs to {pairs_path}")

    return chunks


if __name__ == "__main__":
    run_pipeline()
